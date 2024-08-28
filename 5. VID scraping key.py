import os
import openpyxl
from pip import Document

# Prompt for the base directory containing the subfolders with docx files
base_directory = input("Enter the base directory: ")

# Create or load the Excel workbook
# name = input("Enter the base directory: ")
# base_path = r'C:\Users\Denis.ermakov\OneDrive - Opus Inspection\!1PROJECTS\BIG Migration\\'
# excel_path = os.path.join(base_path, name + ".xlsx")
# if not os.path.exists(excel_path):
#     wb = openpyxl.Workbook()
# else:
#     wb = openpyxl.load_workbook(excel_path)
# sheet = wb.active
#
#
# def get_first_line(file_path):
#     print (file_path)
#     document = Document(file_path)
#     if document.paragraphs:
#         return document.paragraphs[0].text
#     return ""  # Return empty string if no content is found
#
#
# # Start writing from the first row
# row_num = 1
# for folder in os.listdir(base_directory):
#     subfolder_path = os.path.join(base_directory, folder)
#     if os.path.isdir(subfolder_path):  # Check if it's a directory
#         for filename in os.listdir(subfolder_path):
#             if filename.endswith('.docx'):
#                 doc_path = os.path.join(subfolder_path, filename)
#                 first_line = get_first_line(doc_path)
#
#                 # Insert data into columns
#                 sheet.cell(row=row_num, column=1, value=folder)  # Optional: folder name to column A
#                 sheet.cell(row=row_num, column=2, value=os.path.splitext(filename)[0])  # File name without extension
#                 sheet.cell(row=row_num, column=3, value=first_line)  # First line of the document
#                 sheet.cell(row=row_num, column=4, value=doc_path)  # Full file path
#
#                 row_num += 1
#
# wb.save(excel_path)
# print("Data extraction complete.")
# print ("File saved to:" + excel_path)


# Prompt for the base directory containing the subfolders with docx files
name = input("Enter the base directory: ")
base_path = r'C:\Users\Denis.ermakov\OneDrive - Opus Inspection\!1PROJECTS\BIG Migration\\'
excel_path = os.path.join(base_path, name + ".xlsx")

# Create or load the Excel workbook
excel_path = r'C:\Users\Denis.ermakov\OneDrive - Opus Inspection\!1PROJECTS\BIG Migration\your_new_excel_file.xlsx'
if not os.path.exists(excel_path):
    wb = openpyxl.Workbook()
else:
    wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

def get_first_line(file_path):
    print(file_path)
    document = Document(file_path)
    if document.paragraphs:
        return document.paragraphs[0].text
    return ""  # Return empty string if no content is found

def get_primary_keys(file_path):
    document = Document(file_path)
    primary_key = ""
    lane_primary_key = ""
    for paragraph in document.paragraphs:
        if 'Primary Key: ' in paragraph.text:
            primary_key = paragraph.text.split('Primary Key: ')[1]
        if 'Lane Primary Key: ' in paragraph.text:
            lane_primary_key = paragraph.text.split('Lane Primary Key: ')[1]
    return primary_key, lane_primary_key

# Start writing from the first row
row_num = 1
for folder in os.listdir(base_directory):
    subfolder_path = os.path.join(base_directory, folder)
    if os.path.isdir(subfolder_path):  # Check if it's a directory
        for filename in os.listdir(subfolder_path):
            if filename.endswith('.docx'):
                doc_path = os.path.join(subfolder_path, filename)
                first_line = get_first_line(doc_path)
                primary_key, lane_primary_key = get_primary_keys(doc_path)

                # Insert data into columns
                sheet.cell(row=row_num, column=1, value=folder)  # Folder name to column A
                sheet.cell(row=row_num, column=2, value=os.path.splitext(filename)[0])  # File name without extension
                sheet.cell(row=row_num, column=3, value=first_line)  # First line of the document
                sheet.cell(row=row_num, column=4, value=doc_path)  # Full file path
                sheet.cell(row=row_num, column=5, value='Primary Key: ' + primary_key)  # Primary Key
                sheet.cell(row=row_num, column=6, value='Lane Primary Key: ' + lane_primary_key)  # Lane Primary Key

                row_num += 1

wb.save(excel_path)
print("Data extraction complete.")
print("File saved to:" + excel_path)


# ============================================================================================================
# looking up the keys code
# ============================================================================================================
#
# # Prompt for the base directory containing the subfolders with docx files
# base_directory = input("Enter the base directory: ")
#
# # Create or load the Excel workbook
# excel_path = r'C:\Users\Denis.ermakov\OneDrive - Opus Inspection\!1PROJECTS\BIG Migration\your_new_excel_file.xlsx'
# if not os.path.exists(excel_path):
#     wb = openpyxl.Workbook()
# else:
#     wb = openpyxl.load_workbook(excel_path)
# sheet = wb.active
#
# def get_first_line(file_path):
#     print(file_path)
#     document = Document(file_path)
#     if document.paragraphs:
#         return document.paragraphs[0].text
#     return ""  # Return empty string if no content is found
#
# def get_primary_keys(file_path):
#     document = Document(file_path)
#     primary_key = ""
#     lane_primary_key = ""
#     for paragraph in document.paragraphs:
#         if 'Primary Key: ' in paragraph.text:
#             primary_key = paragraph.text.split('Primary Key: ')[1]
#         if 'Lane Primary Key: ' in paragraph.text:
#             lane_primary_key = paragraph.text.split('Lane Primary Key: ')[1]
#     return primary_key, lane_primary_key
#
# # Start writing from the first row
# row_num = 1
# for folder in os.listdir(base_directory):
#     subfolder_path = os.path.join(base_directory, folder)
#     if os.path.isdir(subfolder_path):  # Check if it's a directory
#         for filename in os.listdir(subfolder_path):
#             if filename.endswith('.docx'):
#                 doc_path = os.path.join(subfolder_path, filename)
#                 first_line = get_first_line(doc_path)
#                 primary_key, lane_primary_key = get_primary_keys(doc_path)
#
#                 # Insert data into columns
#                 sheet.cell(row=row_num, column=1, value=folder)  # Folder name to column A
#                 sheet.cell(row=row_num, column=2, value=os.path.splitext(filename)[0])  # File name without extension
#                 sheet.cell(row=row_num, column=3, value=first_line)  # First line of the document
#                 sheet.cell(row=row_num, column=4, value=doc_path)  # Full file path
#                 sheet.cell(row=row_num, column=5, value='Primary Key: ' + primary_key)  # Primary Key
#                 sheet.cell(row=row_num, column=6, value='Lane Primary Key: ' + lane_primary_key)  # Lane Primary Key
#
#                 row_num += 1
#
# wb.save(excel_path)
# print("Data extraction complete.")
# print("File saved to:" + excel_path)
