import os
import openpyxl
from docx import Document

base_directory = input("Enter base directory: ")
states = ["CO", "IN", "MD", "OH"]
print("step 1")

def extract_tables_from_doc(file_path):
    document = Document(file_path)
    tables = []
    for table in document.tables:
        data = []
        for row in table.rows:
            rowData = []
            for cell in row.cells:
                rowData.append(cell.text)
            data.append(rowData)
        tables.append(data)
    return tables

# Create or load the Excel workbook
excel_path = input("Enter location of excel file where you want to save the data: ")
print(excel_path)
if not os.path.exists(excel_path):
    wb = openpyxl.Workbook()
else:
    wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

# Start writing from the first row
row_num = 1
print("step 2")

for state in states:
    test_folder = os.path.join(base_directory, state, "Test")
    print(test_folder)
    print(os.path.exists(test_folder))
    if os.path.exists(test_folder):
        for filename in os.listdir(test_folder):
            if filename.endswith('.docx'):
                doc_path = os.path.join(test_folder, filename)
                tables = extract_tables_from_doc(doc_path)
                print("Step 3")
                for table in tables:
                    for row in table:
                        sheet.cell(row=row_num, column=1, value=state)
                        # CHANGED: Here, only replace .docx to ensure table names don't get an extra 'x'
                        sheet.cell(row=row_num, column=2, value=filename.replace('.docx', ''))
                        for col_num, cell_value in enumerate(row, 3):
                            sheet.cell(row=row_num, column=col_num, value=cell_value)
                        # ADDED: Add the document path as the last column
                        sheet.cell(row=row_num, column=col_num+1, value=doc_path)
                        row_num += 1

wb.save(excel_path)
